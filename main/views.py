from django.shortcuts import render, get_object_or_404
from django.db.models import Sum, Count, F, Max
from .models import Maqola, Kitob, Shogird, Video, Fotogalereya, Xotira
from datetime import datetime, timedelta
import json
import os

def index(request):
    # Bazadan ma'lumotlar (ustoz_haqida kategoriyasini chiqarib tashlash)
    maqolalar_count = Maqola.objects.exclude(category='ustoz_haqida').count()
    kitoblar_count = Kitob.objects.count()
    videolar_count = Video.objects.exclude(category='ustoz_haqida').count()
    shogirdlar_count = Shogird.objects.count()
    
    # Eng so'nggi kitoblar (3 ta)
    latest_books = Kitob.objects.order_by('-year', '-id')[:3]
    
    # Eng so'nggi videolar (3 ta, ustoz_haqida ni chiqarib tashlash)
    latest_videos = Video.objects.exclude(category='ustoz_haqida').order_by('-id')[:3]
    
    # Shogirdlar (4 ta)
    shogirdlar = Shogird.objects.all().order_by('tartib', '-defense_year')[:4]
    
    context = {
        'maqolalar_count': maqolalar_count,
        'kitoblar_count': kitoblar_count,
        'videolar_count': videolar_count,
        'shogirdlar_count': shogirdlar_count,
        'latest_books': latest_books,
        'latest_videos': latest_videos,
        'shogirdlar': shogirdlar,
    }
    
    return render(request, 'index.html', context)

def konkordans(request):
    from .utils import extract_text_from_pdf
    
    query = request.GET.get('q', '').strip()
    category = request.GET.get('category', '')
    author = request.GET.get('author', '')
    limit = int(request.GET.get('limit', 25))
    
    results = []
    total_count = 0
    
    if query:
        # Maqolalardan qidirish
        maqolalar = Maqola.objects.all()
        if category:
            maqolalar = maqolalar.filter(category=category)
        if author:
            maqolalar = maqolalar.filter(author__icontains=author)
            
        for maqola in maqolalar[:50]:  # Maksimal 50 ta maqola
            if len(results) >= limit:
                break
                
            try:
                # PDF'dan sahifa-bo'yicha matnni olish va qidirish
                import PyPDF2
                with open(maqola.pdf_file.path, 'rb') as f:
                    pdf = PyPDF2.PdfReader(f)
                    query_lower = query.lower()
                    for p_idx, page in enumerate(pdf.pages):
                        if len(results) >= limit:
                            break
                        try:
                            page_text = (page.extract_text() or '')
                            page_text_lower = page_text.lower()
                            idx = page_text_lower.find(query_lower)
                            while idx != -1 and len(results) < limit:
                                left_start = max(0, idx - 50)
                                right_end = min(len(page_text), idx + len(query) + 50)

                                left_context = page_text[left_start:idx]
                                keyword = page_text[idx:idx + len(query)]
                                right_context = page_text[idx + len(query):right_end]

                                if left_start > 0:
                                    left_context = '...' + left_context
                                if right_end < len(page_text):
                                    right_context = right_context + '...'

                                results.append({
                                    'left': left_context.strip(),
                                    'keyword': keyword,
                                    'right': right_context.strip(),
                                    'source_title': maqola.title,
                                    'source_author': maqola.author,
                                    'source_year': str(maqola.year),
                                    'source_type': 'Maqola',
                                    'source_id': maqola.id,
                                    'page': p_idx + 1,
                                })
                                total_count += 1

                                idx = page_text_lower.find(query_lower, idx + 1)
                        except Exception as e:
                            print(f"Page extract error (Maqola {maqola.id} page {p_idx}): {e}")
                            continue
            except Exception as e:
                print(f"Xato (Maqola {maqola.id}): {e}")
                continue
        
        # Kitoblardan qidirish
        if len(results) < limit:
            kitoblar = Kitob.objects.all()
            if author:
                kitoblar = kitoblar.filter(author__icontains=author)
                
            for kitob in kitoblar[:50]:
                if len(results) >= limit:
                    break
                    
                try:
                    # Page-wise search for books
                    import PyPDF2
                    with open(kitob.pdf_file.path, 'rb') as f:
                        pdf = PyPDF2.PdfReader(f)
                        query_lower = query.lower()
                        for p_idx, page in enumerate(pdf.pages):
                            if len(results) >= limit:
                                break
                            try:
                                page_text = (page.extract_text() or '')
                                page_text_lower = page_text.lower()
                                idx = page_text_lower.find(query_lower)
                                while idx != -1 and len(results) < limit:
                                    left_start = max(0, idx - 50)
                                    right_end = min(len(page_text), idx + len(query) + 50)

                                    left_context = page_text[left_start:idx]
                                    keyword = page_text[idx:idx + len(query)]
                                    right_context = page_text[idx + len(query):right_end]

                                    if left_start > 0:
                                        left_context = '...' + left_context
                                    if right_end < len(page_text):
                                        right_context = right_context + '...'

                                    results.append({
                                        'left': left_context.strip(),
                                        'keyword': keyword,
                                        'right': right_context.strip(),
                                        'source_title': kitob.title,
                                        'source_author': kitob.author,
                                        'source_year': str(kitob.year),
                                        'source_type': 'Kitob',
                                        'source_id': kitob.id,
                                        'page': p_idx + 1,
                                    })
                                    total_count += 1

                                    idx = page_text_lower.find(query_lower, idx + 1)
                            except Exception as e:
                                print(f"Page extract error (Kitob {kitob.id} page {p_idx}): {e}")
                                continue
                except Exception as e:
                    print(f"Xato (Kitob {kitob.id}): {e}")
                    continue
    
    # Kategoriyalar va mualliflar ro'yxatini olish
    categories = [choice[1] for choice in Maqola.CATEGORY_CHOICES]
    authors = list(set(
        [m.author for m in Maqola.objects.all() if m.author] + 
        [k.author for k in Kitob.objects.all() if k.author]
    ))
    
    context = {
        'query': query,
        'results': results[:limit],
        'total_count': total_count,
        'categories': categories,
        'authors': sorted(authors),
        'selected_category': category,
        'selected_author': author,
        'limit': limit
    }
    
    return render(request, 'konkordans.html', context)

def kengaytirilgan(request):
    from .utils import extract_text_from_pdf
    import re as regex_module
    
    query = request.GET.get('q', '').strip()
    operator = request.GET.get('operator', 'contains')
    include = request.GET.get('include', '').strip()
    exclude = request.GET.get('exclude', '').strip()
    author = request.GET.get('author', '')
    category = request.GET.get('genre', '')
    context_before = int(request.GET.get('context_before', 5))
    context_after = int(request.GET.get('context_after', 5))
    case_sensitive = request.GET.get('case_sensitive') == 'on'
    whole_word = request.GET.get('whole_word') == 'on'
    use_regex = request.GET.get('regex') == 'on'
    
    results = []
    total_count = 0
    
    if query:
        # Maqolalardan qidirish
        maqolalar = Maqola.objects.all()
        if category:
            maqolalar = maqolalar.filter(category=category)
        if author:
            maqolalar = maqolalar.filter(author__icontains=author)
            
        for maqola in maqolalar[:50]:
            if len(results) >= 100:
                break
                
            try:
                text = extract_text_from_pdf(maqola.pdf_file.path)
                
                # Qidiruv patterni yaratish
                if use_regex:
                    pattern = query
                else:
                    if whole_word:
                        pattern = r'\b' + regex_module.escape(query) + r'\b'
                    else:
                        pattern = regex_module.escape(query)
                    
                    if operator == 'starts':
                        pattern = r'\b' + pattern
                    elif operator == 'ends':
                        pattern = pattern + r'\b'
                
                flags = 0 if case_sensitive else regex_module.IGNORECASE
                
                # Qidirish
                for match in regex_module.finditer(pattern, text, flags):
                    if len(results) >= 100:
                        break
                    
                    start = match.start()
                    end = match.end()
                    
                    # Kontekstni olish (so'zlar soni bo'yicha)
                    words_before = text[:start].split()[-context_before:]
                    words_after = text[end:].split()[:context_after]
                    
                    left_context = ' '.join(words_before)
                    keyword = match.group()
                    right_context = ' '.join(words_after)
                    
                    # Include/exclude filterlarini tekshirish
                    full_context = left_context + ' ' + keyword + ' ' + right_context
                    
                    if include:
                        include_words = [w.strip() for w in include.split(',')]
                        if not any(word.lower() in full_context.lower() for word in include_words):
                            continue
                    
                    if exclude:
                        exclude_words = [w.strip() for w in exclude.split(',')]
                        if any(word.lower() in full_context.lower() for word in exclude_words):
                            continue
                    
                    results.append({
                        'left': '...' + left_context if words_before else left_context,
                        'keyword': keyword,
                        'right': right_context + '...' if words_after else right_context,
                        'source_title': maqola.title,
                        'source_author': maqola.author,
                        'source_year': str(maqola.year),
                        'source_type': 'Maqola',
                        'source_id': maqola.id,
                        'category': maqola.get_category_display()
                    })
                    total_count += 1
                    
            except Exception as e:
                print(f"Xato (Maqola {maqola.id}): {e}")
                continue
        
        # Kitoblardan qidirish
        if len(results) < 100:
            kitoblar = Kitob.objects.all()
            if author:
                kitoblar = kitoblar.filter(author__icontains=author)
                
            for kitob in kitoblar[:50]:
                if len(results) >= 100:
                    break
                    
                try:
                    text = extract_text_from_pdf(kitob.pdf_file.path)
                    
                    if use_regex:
                        pattern = query
                    else:
                        if whole_word:
                            pattern = r'\b' + regex_module.escape(query) + r'\b'
                        else:
                            pattern = regex_module.escape(query)
                        
                        if operator == 'starts':
                            pattern = r'\b' + pattern
                        elif operator == 'ends':
                            pattern = pattern + r'\b'
                    
                    flags = 0 if case_sensitive else regex_module.IGNORECASE
                    
                    for match in regex_module.finditer(pattern, text, flags):
                        if len(results) >= 100:
                            break
                        
                        start = match.start()
                        end = match.end()
                        
                        words_before = text[:start].split()[-context_before:]
                        words_after = text[end:].split()[:context_after]
                        
                        left_context = ' '.join(words_before)
                        keyword = match.group()
                        right_context = ' '.join(words_after)
                        
                        full_context = left_context + ' ' + keyword + ' ' + right_context
                        
                        if include:
                            include_words = [w.strip() for w in include.split(',')]
                            if not any(word.lower() in full_context.lower() for word in include_words):
                                continue
                        
                        if exclude:
                            exclude_words = [w.strip() for w in exclude.split(',')]
                            if any(word.lower() in full_context.lower() for word in exclude_words):
                                continue
                        
                        results.append({
                            'left': '...' + left_context if words_before else left_context,
                            'keyword': keyword,
                            'right': right_context + '...' if words_after else right_context,
                            'source_title': kitob.title,
                            'source_author': kitob.author,
                            'source_year': str(kitob.year),
                            'source_type': 'Kitob',
                            'source_id': kitob.id
                        })
                        total_count += 1
                        
                except Exception as e:
                    print(f"Xato (Kitob {kitob.id}): {e}")
                    continue
    
    # Mualliflar va kategoriyalar ro'yxatini olish
    categories = [choice[1] for choice in Maqola.CATEGORY_CHOICES]
    authors = list(set(
        [m.author for m in Maqola.objects.all() if m.author] + 
        [k.author for k in Kitob.objects.all() if k.author]
    ))
    
    context = {
        'query': query,
        'results': results,
        'total_count': total_count,
        'categories': categories,
        'authors': sorted(authors),
        'selected_author': author,
        'selected_genre': category,
        'operator': operator,
        'include': include,
        'exclude': exclude,
        'context_before': context_before,
        'context_after': context_after,
        'case_sensitive': case_sensitive,
        'whole_word': whole_word,
        'use_regex': use_regex
    }
    
    return render(request, 'kengaytirilgan.html', context)

def tarjimai_hol(request):
    return render(request, 'tarjimai-hol.html')


def korpus_haqida(request):
    # Korpus statistikasi
    stats = {
        'total_words': '15,000,000',
        'documents': '25,000',
        'authors': '350',
        'years': '1920-2024',
        'genres': '12'
    }
    
    # Rahbar
    leader = {
        'name': 'Toirova Guli Ibragimovna',
        'title': 'Loyiha rahbari',
        'degree': 'Filologiya fanlari doktori (DSc), professor',
        'image': '/static/images/rahbar.jpg',
        'bio': ' “Ergash Qilichev mualliflik korpusi”ni yaratish g‘oyasi muallifi sifatida mazkur loyihaning ilmiy konsepsiyasi, metodologik asoslari va modulli arxitekturasini ishlab chiqqan. Uning tashabbusi bilan korpus muallifga yo‘naltirilgan (author-centered) raqamli platforma sifatida loyihalashtirilib, ilmiy merosni raqamlashtirish, strukturlashtirish va intellektual tahlil qilish imkoniyatlari yaratilgan.',
        'email': 'tugulijon@mail.ru',
        'phone': '+998902993417',
        'address': 'Buxoro sh. M.Iqbol ko\'chasi 11-uy',
        'work_time': '8:30dan 17:00gacha'
    }
    
    # Shogirdlar (tuzuvchilar)
    contributors = [
        {
            'name': 'Dr. Nodira Rahimova',
            'role': 'Morfologiya bo\'limi rahbari',
            'specialty': 'Morfologik tahlil va annotatsiya',
            'image': 'https://images.unsplash.com/photo-1494790108377-be9c29b29330?w=400',
            'contribution': 'Morfologik teg tizimini ishlab chiqish, 5 million so\'zni annotatsiya qilish'
        },
        {
            'name': 'Dr. Jasur Karimov',
            'role': 'Sintaksis bo\'limi rahbari',
            'specialty': 'Sintaktik tahlil va parsing',
            'image': 'https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?w=400',
            'contribution': 'Sintaktik parser yaratish, gap tuzilmalarini tahlil qilish tizimlari'
        },
        {
            'name': 'Dilnoza Abdullayeva',
            'role': 'Matn to\'plash va tayyorlash',
            'specialty': 'Korpus arxitekturasi',
            'image': 'https://images.unsplash.com/photo-1438761681033-6461ffad8d80?w=400',
            'contribution': '15,000 ta matnni raqamlashtirish, korpus strukturasini loyihalash'
        },
        {
            'name': 'Bobur Rahmonov',
            'role': 'Texnik ishlar bo\'limi',
            'specialty': 'Dasturiy ta\'minot',
            'image': 'https://images.unsplash.com/photo-1500648767791-00dcc994a43e?w=400',
            'contribution': 'Qidiruv tizimini yaratish, web interfeys va API ishlab chiqish'
        },
        {
            'name': 'Malika Turgunova',
            'role': 'Leksikografiya bo\'limi',
            'specialty': 'Lug\'at tuzish',
            'image': 'https://images.unsplash.com/photo-1487412720507-e7ab37603c6f?w=400',
            'contribution': 'Chastota lug\'ati, kollokatsiyalar bazasi va semantik teg tizimi'
        }
    ]
    
    context = {
        'stats': stats,
        'leader': leader,
        'contributors': contributors
    }
    return render(request, 'korpus-haqida.html', context)

def statistika(request):
    import json
    import os
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count, Q, Max
    from django.conf import settings
    from .utils import extract_text_from_pdf
    from collections import Counter
    import re as regex_module
    
    # JSON cache fayl yo'li
    cache_dir = os.path.join(settings.BASE_DIR, 'cache')
    if not os.path.exists(cache_dir):
        os.makedirs(cache_dir)
    cache_file = os.path.join(cache_dir, 'statistika_cache.json')
    
    # Cache'ni tekshirish funksiyasi
    def should_update_cache():
        # Cache fayl mavjud emasmi?
        if not os.path.exists(cache_file):
            return True
        
        # Cache faylning yaratilgan vaqti
        cache_time = datetime.fromtimestamp(os.path.getmtime(cache_file))
        
        # Ma'lumotlar bazasidagi eng oxirgi yangilanish vaqti
        latest_maqola = Maqola.objects.aggregate(Max('updated_at'))['updated_at__max']
        latest_kitob = Kitob.objects.aggregate(Max('updated_at'))['updated_at__max']
        latest_video = Video.objects.aggregate(Max('updated_at'))['updated_at__max']
        
        latest_updates = [latest_maqola, latest_kitob, latest_video]
        latest_updates = [dt for dt in latest_updates if dt is not None]
        
        if latest_updates:
            latest_db_update = max(latest_updates)
            # Agar ma'lumotlar bazasi cache'dan keyinroq yangilangan bo'lsa
            if latest_db_update.replace(tzinfo=None) > cache_time:
                return True
        
        # Cache 24 soatdan eski bo'lsa
        if datetime.now() - cache_time > timedelta(hours=24):
            return True
        
        return False
    
    # Cache'dan ma'lumot o'qish
    def load_from_cache():
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Cache o'qishda xato: {e}")
            return None
    
    # Cache'ga saqlash
    def save_to_cache(data):
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"Statistika cache yangilandi: {datetime.now()}")
        except Exception as e:
            print(f"Cache saqlashda xato: {e}")
    
    # Agar cache'ni yangilash kerak bo'lmasa, mavjud cache'dan o'qish
    if not should_update_cache():
        cached_data = load_from_cache()
        if cached_data:
            print("Cache'dan statistika yuklandi")
            return render(request, 'statistika.html', cached_data)
    
    print("Statistika qayta hisoblanmoqda...")
    
    # Asosiy statistika
    maqolalar = Maqola.objects.all()
    kitoblar = Kitob.objects.all()
    videolar = Video.objects.all()
    
    total_maqolalar = maqolalar.count()
    total_kitoblar = kitoblar.count()
    total_videolar = videolar.count()
    total_documents = total_maqolalar + total_kitoblar
    
    # So'zlar va statistika
    total_words = maqolalar.aggregate(Sum('total_words'))['total_words__sum'] or 0
    total_words += kitoblar.aggregate(Sum('total_words'))['total_words__sum'] or 0
    
    total_sentences = maqolalar.aggregate(Sum('total_sentences'))['total_sentences__sum'] or 0
    total_sentences += kitoblar.aggregate(Sum('total_sentences'))['total_sentences__sum'] or 0
    
    total_paragraphs = maqolalar.aggregate(Sum('total_paragraphs'))['total_paragraphs__sum'] or 0
    total_paragraphs += kitoblar.aggregate(Sum('total_paragraphs'))['total_paragraphs__sum'] or 0
    
    total_characters = maqolalar.aggregate(Sum('total_characters'))['total_characters__sum'] or 0
    total_characters += kitoblar.aggregate(Sum('total_characters'))['total_characters__sum'] or 0

    # Belgilar (bo'sh joysiz) - character counts without spaces
    total_characters_no_spaces = maqolalar.aggregate(Sum('total_characters_no_spaces'))['total_characters_no_spaces__sum'] or 0
    total_characters_no_spaces += kitoblar.aggregate(Sum('total_characters_no_spaces'))['total_characters_no_spaces__sum'] or 0
    
    unique_words = maqolalar.aggregate(Sum('unique_words'))['unique_words__sum'] or 0
    unique_words += kitoblar.aggregate(Sum('unique_words'))['unique_words__sum'] or 0
    
    # Mualliflar soni
    maqola_authors = set([m.author for m in maqolalar if m.author])
    kitob_authors = set([k.author for k in kitoblar if k.author])
    total_authors = len(maqola_authors | kitob_authors)
    
    # Kategoriya bo'yicha (Maqolalar + Kitoblar)
    categories = []
    for choice in Maqola.CATEGORY_CHOICES:
        if choice[0] != 'ustoz_haqida':  # Ustoz haqida kategoriyasini chiqarib tashlash
            cat_maqolalar = maqolalar.filter(category=choice[0])
            cat_count = cat_maqolalar.aggregate(Sum('total_words'))['total_words__sum'] or 0
            
            # Kitoblar uchun kategoriya yo'q, lekin umumiy statistikaga qo'shamiz
            if cat_count > 0:
                percent = int((cat_count / total_words * 100)) if total_words > 0 else 0
                categories.append({
                    'name': choice[1],
                    'count': cat_count,
                    'percent': percent
                })
    
    # Yil bo'yicha
    year_ranges = [
        ('1960-2000', 1960, 2000),
        ('2001-2005', 2001, 2005),
        ('2006-2025', 2006, 2025),
    ]
    
    years = []
    for period, start, end in year_ranges:
        period_maqolalar = maqolalar.filter(year__gte=start, year__lte=end)
        period_kitoblar = kitoblar.filter(year__gte=start, year__lte=end)
        
        period_words = period_maqolalar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        period_words += period_kitoblar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        
        if period_words > 0:
            percent = int((period_words / total_words * 100)) if total_words > 0 else 0
            years.append({
                'period': period,
                'count': period_words,
                'percent': percent
            })
    
    # Eng ko'p uchraydigan so'zlarni topish
    word_counter = Counter()
    
    # Barcha PDF'lardan so'zlarni hisoblash
    all_maqolalar = list(maqolalar)
    all_kitoblar = list(kitoblar)
    
    for maqola in all_maqolalar:
        try:
            text = extract_text_from_pdf(maqola.pdf_file.path)
            # So'zlarni ajratish va tozalash
            words = regex_module.findall(r'\b[a-zA-Zа-яА-ЯўғҳқўҒҲҚЎ\']+\b', text.lower())
            word_counter.update(words)
        except Exception as e:
            print(f"Xato (Maqola {maqola.id}): {e}")
            continue
    
    for kitob in all_kitoblar:
        try:
            text = extract_text_from_pdf(kitob.pdf_file.path)
            words = regex_module.findall(r'\b[a-zA-Zа-яА-ЯўғҳқўҒҲҚЎ\']+\b', text.lower())
            word_counter.update(words)
        except Exception as e:
            print(f"Xato (Kitob {kitob.id}): {e}")
            continue
    
    # Kamida 75 marta uchraydigan so'zlar
    top_words = [
        {'word': word, 'count': count}
        for word, count in word_counter.most_common()
        if count >= 75]

    
    # Eng ko'p asarlarga ega mualliflar
    author_stats = []
    for author in list(maqola_authors | kitob_authors)[:10]:
        author_maqolalar = maqolalar.filter(author__icontains=author)
        author_kitoblar = kitoblar.filter(author__icontains=author)
        
        works = author_maqolalar.count() + author_kitoblar.count()
        words = author_maqolalar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        words += author_kitoblar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        
        if works > 0:
            author_stats.append({
                'name': author,
                'works': works,
                'words': words
            })
    
    # Eng ko'p asarlarga ega bo'yicha saralash
    author_stats.sort(key=lambda x: x['works'], reverse=True)
    
    context = {
        'total_words': total_words,
        'total_documents': total_documents,
        'total_sentences': total_sentences,
        'total_paragraphs': total_paragraphs,
        'total_characters': total_characters,
        'total_characters_no_spaces': total_characters_no_spaces,
        'unique_words': unique_words,
        'total_authors': total_authors,
        'total_maqolalar': total_maqolalar,
        'total_kitoblar': total_kitoblar,
        'total_videolar': total_videolar,
        'categories': categories,
        'years': years,
        'top_words': top_words,
        'top_authors': author_stats[:10],
    }
    
    # Ma'lumotlarni keshga saqlash
    save_to_cache(context)
    print("Statistika ma'lumotlari qayta hisoblandi va yangi keshga saqlandi")
    
    return render(request, 'statistika.html', context)

def statistika_export(request):
    """Statistikani Excel fayliga eksport qilish"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    from django.db.models import Sum, Count
    from collections import Counter
    import re as regex_module
    from .utils import extract_text_from_pdf
    
    # Ma'lumotlarni to'plash (statistika view'dagi kod)
    maqolalar = Maqola.objects.all()
    kitoblar = Kitob.objects.all()
    videolar = Video.objects.all()
    
    total_maqolalar = maqolalar.count()
    total_kitoblar = kitoblar.count()
    total_videolar = videolar.count()
    total_documents = total_maqolalar + total_kitoblar
    
    total_words = maqolalar.aggregate(Sum('total_words'))['total_words__sum'] or 0
    total_words += kitoblar.aggregate(Sum('total_words'))['total_words__sum'] or 0
    
    total_sentences = maqolalar.aggregate(Sum('total_sentences'))['total_sentences__sum'] or 0
    total_sentences += kitoblar.aggregate(Sum('total_sentences'))['total_sentences__sum'] or 0
    
    total_paragraphs = maqolalar.aggregate(Sum('total_paragraphs'))['total_paragraphs__sum'] or 0
    total_paragraphs += kitoblar.aggregate(Sum('total_paragraphs'))['total_paragraphs__sum'] or 0
    
    total_characters = maqolalar.aggregate(Sum('total_characters'))['total_characters__sum'] or 0
    total_characters += kitoblar.aggregate(Sum('total_characters'))['total_characters__sum'] or 0
    
    unique_words = maqolalar.aggregate(Sum('unique_words'))['unique_words__sum'] or 0
    unique_words += kitoblar.aggregate(Sum('unique_words'))['unique_words__sum'] or 0
    
    maqola_authors = set([m.author for m in maqolalar if m.author])
    kitob_authors = set([k.author for k in kitoblar if k.author])
    total_authors = len(maqola_authors | kitob_authors)
    
    # Kategoriyalar
    categories = []
    for choice in Maqola.CATEGORY_CHOICES:
        cat_maqolalar = maqolalar.filter(category=choice[0])
        cat_count = cat_maqolalar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        if cat_count > 0:
            percent = int((cat_count / total_words * 100)) if total_words > 0 else 0
            categories.append({
                'name': choice[1],
                'count': cat_count,
                'percent': percent
            })
    
    # Yillar
    year_ranges = [
        ('1920-1940', 1920, 1940),
        ('1941-1960', 1941, 1960),
        ('1961-1980', 1961, 1980),
        ('1981-2000', 1981, 2000),
        ('2001-2024', 2001, 2024)
    ]
    
    years = []
    for period, start, end in year_ranges:
        period_maqolalar = maqolalar.filter(year__gte=start, year__lte=end)
        period_kitoblar = kitoblar.filter(year__gte=start, year__lte=end)
        
        period_words = period_maqolalar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        period_words += period_kitoblar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        
        if period_words > 0:
            percent = int((period_words / total_words * 100)) if total_words > 0 else 0
            years.append({
                'period': period,
                'count': period_words,
                'percent': percent
            })
    
    # Mualliflar
    author_stats = []
    for author in list(maqola_authors | kitob_authors)[:20]:
        author_maqolalar = maqolalar.filter(author__icontains=author)
        author_kitoblar = kitoblar.filter(author__icontains=author)
        
        works = author_maqolalar.count() + author_kitoblar.count()
        words = author_maqolalar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        words += author_kitoblar.aggregate(Sum('total_words'))['total_words__sum'] or 0
        
        if works > 0:
            author_stats.append({
                'name': author,
                'works': works,
                'words': words
            })
    
    author_stats.sort(key=lambda x: x['works'], reverse=True)
    
    # Excel fayl yaratish
    wb = Workbook()
    
    # Stil sozlamalari
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="1e3a8a")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Asosiy statistika
    ws1 = wb.active
    ws1.title = "Asosiy Ko'rsatkichlar"
    
    ws1['A1'] = "O'ZBEK TILI KORPUS STATISTIKASI"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:B1')
    
    ws1['A2'] = f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1.merge_cells('A2:B2')
    
    ws1['A4'] = "Ko'rsatkich"
    ws1['B4'] = "Qiymat"
    ws1['A4'].font = header_font
    ws1['B4'].font = header_font
    ws1['A4'].fill = header_fill
    ws1['B4'].fill = header_fill
    
    data = [
        ["Jami so'zlar", total_words],
        ["Jami hujjatlar", total_documents],
        ["Jami gaplar", total_sentences],
        ["Jami paragraflar", total_paragraphs],
        ["Jami belgilar", total_characters],
        ["Noyob so'zlar", unique_words],
        ["Mualliflar soni", total_authors],
        ["Maqolalar", total_maqolalar],
        ["Kitoblar", total_kitoblar],
        ["Videolar", total_videolar],
    ]
    
    for idx, (label, value) in enumerate(data, start=5):
        ws1[f'A{idx}'] = label
        ws1[f'B{idx}'] = value
        ws1[f'A{idx}'].border = border
        ws1[f'B{idx}'].border = border
    
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    
    # 2. Janr bo'yicha
    ws2 = wb.create_sheet("Janrlar")
    ws2['A1'] = "JANRLAR BO'YICHA TAQSIMOT"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:C1')
    
    ws2['A3'] = "Janr"
    ws2['B3'] = "So'zlar soni"
    ws2['C3'] = "Foiz"
    for cell in ['A3', 'B3', 'C3']:
        ws2[cell].font = header_font
        ws2[cell].fill = header_fill
    
    for idx, cat in enumerate(categories, start=4):
        ws2[f'A{idx}'] = cat['name']
        ws2[f'B{idx}'] = cat['count']
        ws2[f'C{idx}'] = f"{cat['percent']}%"
        for cell in [f'A{idx}', f'B{idx}', f'C{idx}']:
            ws2[cell].border = border
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 15
    
    # 3. Yillar bo'yicha
    ws3 = wb.create_sheet("Yillar")
    ws3['A1'] = "YILLAR BO'YICHA TAQSIMOT"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:C1')
    
    ws3['A3'] = "Davr"
    ws3['B3'] = "So'zlar soni"
    ws3['C3'] = "Foiz"
    for cell in ['A3', 'B3', 'C3']:
        ws3[cell].font = header_font
        ws3[cell].fill = header_fill
    
    for idx, year in enumerate(years, start=4):
        ws3[f'A{idx}'] = year['period']
        ws3[f'B{idx}'] = year['count']
        ws3[f'C{idx}'] = f"{year['percent']}%"
        for cell in [f'A{idx}', f'B{idx}', f'C{idx}']:
            ws3[cell].border = border
    
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 15
    
    # 4. Mualliflar
    ws4 = wb.create_sheet("Mualliflar")
    ws4['A1'] = "ENG KO'P ASARLARGA EGA MUALLIFLAR"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:C1')
    
    ws4['A3'] = "Muallif"
    ws4['B3'] = "Asarlar soni"
    ws4['C3'] = "So'zlar soni"
    for cell in ['A3', 'B3', 'C3']:
        ws4[cell].font = header_font
        ws4[cell].fill = header_fill
    
    for idx, author in enumerate(author_stats[:20], start=4):
        ws4[f'A{idx}'] = author['name']
        ws4[f'B{idx}'] = author['works']
        ws4[f'C{idx}'] = author['words']
        for cell in [f'A{idx}', f'B{idx}', f'C{idx}']:
            ws4[cell].border = border
    
    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 20
    
    # Faylni saqlash
    filename = f"Statistics_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def videolar(request):
    # Video va fotogalereya ma'lumotlarini bazadan olish (ustoz_haqida ni chiqarib tashlash)
    videos = Video.objects.exclude(category='ustoz_haqida').order_by('-published_date')
    photos = Fotogalereya.objects.all().order_by('-date')
    
    context = {
        'videos': videos,
        'photos': photos
    }
    return render(request, 'videolar.html', context)

def video_detail(request, id):
    """Video detail sahifasi va ko'rishlar sonini oshirish"""
    from django.shortcuts import get_object_or_404
    from django.db.models import F
    
    video = get_object_or_404(Video, id=id)
    
    # Ko'rishlar sonini oshirish
    Video.objects.filter(id=id).update(views=F('views') + 1)
    video.refresh_from_db()
    
    return render(request, 'video-detail.html', {'video': video})

def video_track(request, id):
    """Video ko'rilganini track qilish (80% yoki to'liq ko'rilganda)"""
    from django.shortcuts import get_object_or_404
    from django.http import JsonResponse
    from django.db.models import F
    from django.views.decorators.csrf import csrf_exempt
    import json
    
    if request.method == 'POST':
        try:
            video = get_object_or_404(Video, id=id)
            
            # Ko'rishlar sonini oshirish
            Video.objects.filter(id=id).update(views=F('views') + 1)
            video.refresh_from_db()
            
            return JsonResponse({
                'status': 'success',
                'message': 'Video ko\'rildi deb belgilandi',
                'views': video.views
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
    
    return JsonResponse({
        'status': 'error',
        'message': 'Faqat POST so\'rov qabul qilinadi'
    }, status=405)

def shogirdlar(request):
    # Bazadan barcha shogirdlarni olish
    shogirdlar_list = Shogird.objects.all().order_by('tartib', '-defense_year', 'full_name')
    
    context = {
        'shogirdlar': shogirdlar_list
    }
    return render(request, 'shogirdlar.html', context)

def shogird_detail(request, id):
    from django.shortcuts import get_object_or_404
    from django.db.models import F
    
    shogird = get_object_or_404(Shogird, id=id)
    
    # Ko'rishlar sonini oshirish
    Shogird.objects.filter(id=id).update(views=F('views') + 1)
    shogird.refresh_from_db()
    
    return render(request, 'shogird-detail.html', {'shogird': shogird})

def maqolalar(request):
    # Kategoriya bo'yicha filtrlash
    category = request.GET.get('category', 'all')
    page = request.GET.get('page', 1)
    
    # Database query optimization - select only needed fields
    base_fields = ['id', 'title', 'author', 'year', 'category', 'journal', 'total_pages', 'total_words', 'views', 'tartib', 'created_at']
    
    if category == 'all' or not category:
        maqolalar_list = Maqola.objects.exclude(category='ustoz_haqida').only(*base_fields).order_by('tartib', '-year', '-created_at')
    else:
        maqolalar_list = Maqola.objects.filter(category=category).exclude(category='ustoz_haqida').only(*base_fields).order_by('tartib', '-year', '-created_at')
    
    # Pagination qo'shish - har sahifada 12 ta maqola  
    from django.core.paginator import Paginator
    paginator = Paginator(maqolalar_list, 12)
    
    try:
        maqolalar = paginator.page(page)
    except:
        maqolalar = paginator.page(1)
    
    # Har bir kategoriyada nechta maqola borligini hisoblash (ustoz_haqida ni chiqarib tashlash)
    categories_with_count = []
    for choice in Maqola.CATEGORY_CHOICES:
        if choice[0] != 'ustoz_haqida':
            count = Maqola.objects.filter(category=choice[0]).count()
            if count > 0 or choice[0] == Maqola.CATEGORY_CHOICES[0][0]:
                categories_with_count.append({
                    'key': choice[0],
                    'name': choice[1],
                    'count': count
                })
    
    # "Barchasi" kategoriyasini birinchiga qo'shish
    categories_with_count.insert(0, {
        'key': 'all',
        'name': 'Barchasi',
        'count': Maqola.objects.exclude(category='ustoz_haqida').count()
    })
    
    # Faqat maqola bor kategoriyalarni ko'rsatish
    categories = [cat for cat in categories_with_count if cat['count'] > 0]
    
    # Umumiy statistikani hisoblash
    stats = Maqola.objects.exclude(category='ustoz_haqida').aggregate(
        total_articles=Count('id'),
        total_pages=Sum('total_pages'),
        total_words=Sum('total_words'),
        total_views=Sum('views')
    )
    
    context = {
        'maqolalar': maqolalar,
        'categories': categories,
        'current_category': category,
        'stats': stats,
        'paginator': paginator,
        'page_obj': maqolalar,
        'is_paginated': paginator.num_pages > 1,
    }
    return render(request, 'maqolalar.html', context)

def maqola_detail(request, id):
    # Database'dan maqola olish
    maqola = get_object_or_404(Maqola, id=id)
    
    # Ko'rishlar sonini oshirish
    Maqola.objects.filter(id=id).update(views=maqola.views + 1)
    
    context = {
        'maqola': maqola
    }
    return render(request, 'maqola-detail.html', context)

def kitoblar(request):
    # Bazadan barcha kitoblarni olish
    kitoblar_list = Kitob.objects.all().order_by('-year', '-created_at')
    
    # Umumiy statistikani hisoblash
    stats = Kitob.objects.aggregate(
        total_books=Count('id'),
        total_pages=Sum('total_pages'),
        total_words=Sum('total_words'),
        total_views=Sum('views')
    )
    
    context = {
        'kitoblar': kitoblar_list,
        'stats': stats
    }
    return render(request, 'kitoblar.html', context)

def kitob_detail(request, id):
    # Bazadan kitobni olish
    kitob = get_object_or_404(Kitob, id=id)
    
    # Ko'rishlar sonini oshirish
    Kitob.objects.filter(id=id).update(views=kitob.views + 1)
    
    context = {
        'kitob': kitob
    }
    return render(request, 'kitob-detail.html', context)

def ustoz_haqida(request):
    """Ustoz haqida sahifasi - faqat ustoz_haqida kategoriyasidagi maqola va videolar"""
    
    # Faqat ustoz_haqida kategoriyasidagi maqola va videolarni olish
    maqolalar_list = Maqola.objects.filter(category='ustoz_haqida').order_by('-year', '-created_at')
    videos_list = Video.objects.filter(category='ustoz_haqida').order_by('-published_date')
    
    # Statistika
    maqola_stats = {
        'total': maqolalar_list.count(),
        'total_views': maqolalar_list.aggregate(Sum('views'))['views__sum'] or 0
    }
    
    video_stats = {
        'total': videos_list.count(),
        'total_views': videos_list.aggregate(Sum('views'))['views__sum'] or 0
    }
    
    context = {
        'maqolalar': maqolalar_list,
        'videos': videos_list,
        'maqola_stats': maqola_stats,
        'video_stats': video_stats
    }
    return render(request, 'ustoz-haqida.html', context)


def xotiralar(request):
    """Xotiralar sahifasi - sodda tartibda"""
    
    # Barcha xotiralarni yangi yaratilgan tartibda olish
    xotiralar_list = Xotira.objects.all().order_by('-id')
    
    context = {
        'xotiralar': xotiralar_list,
    }
    return render(request, 'xotiralar.html', context)


def xotira_detail(request, id):
    """Xotira detali sahifasi"""
    
    # Xotirani olish
    xotira = get_object_or_404(Xotira, id=id)
    
    # Ko'rishlar sonini oshirish
    Xotira.objects.filter(id=id).update(views=F('views') + 1)
    xotira.refresh_from_db()
    
    # Teglarni ajratish
    tags = xotira.tag_list
    
    # Keyingi va oldingi xotiralar
    next_memory = Xotira.objects.filter(year__gt=xotira.year).order_by('year', 'id').first()
    prev_memory = Xotira.objects.filter(year__lt=xotira.year).order_by('-year', '-id').first()
    
    # Agar bir xil yilda bo'lsa
    if not next_memory:
        next_memory = Xotira.objects.filter(year=xotira.year, id__gt=xotira.id).order_by('id').first()
    if not prev_memory:
        prev_memory = Xotira.objects.filter(year=xotira.year, id__lt=xotira.id).order_by('-id').first()
    
    context = {
        'xotira': xotira,
        'tags': tags,
        'next_memory': next_memory,
        'prev_memory': prev_memory
    }
    return render(request, 'xotira-detail.html', context)
